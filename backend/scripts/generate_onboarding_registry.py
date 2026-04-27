#!/usr/bin/env python3
"""Re-extract the 114-question salesbook onboarding registry from the canonical
Google Sheet (`1HGSlYMtxE9...`). /Oliviercontribution.

Usage (from repo root):
    python3 backend/scripts/generate_onboarding_registry.py

Outputs (overwrites in place):
    backend/src/hello_sales_backend/modules/salesbook/domain/_onboarding_registry.json
    backend/src/hello_sales_backend/modules/salesbook/domain/onboarding_registry.py
"""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

SHEET_ID = "1HGSlYMtxE9tbk15198wkg-U7n85vdO4u2oNCZP6B2Pw"
SHEET_URL = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"

REPO_ROOT = Path(__file__).resolve().parents[2]
DOMAIN_DIR = REPO_ROOT / "backend/src/hello_sales_backend/modules/salesbook/domain"
DATA_OUT = DOMAIN_DIR / "_onboarding_registry.json"
PY_OUT = DOMAIN_DIR / "onboarding_registry.py"

TMP_XLSX = Path("/tmp/hs-onboarding.xlsx")


def fetch_sheet() -> None:
    print(f"[generate-onboarding-registry] downloading {SHEET_URL}")
    subprocess.check_call(["curl", "-sL", "-o", str(TMP_XLSX), SHEET_URL])


def extract() -> list[dict]:
    try:
        import openpyxl  # type: ignore[import-untyped]
    except ImportError:
        print("openpyxl required: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(TMP_XLSX, data_only=True)

    def _extract(sheet: str, *, phase: int, n_col: int, sec_col: int,
                 sub_col: int | None, q_col: int, t_col: int, ex_col: int,
                 key_col: int, header_row: int) -> list[dict]:
        ws = wb[sheet]
        out: list[dict] = []
        for r in range(header_row + 1, ws.max_row + 1):
            row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            if all(v is None for v in row):
                continue
            if not isinstance(row[n_col], (int, float)):
                continue
            if not row[key_col]:
                continue
            out.append({
                "phase": phase,
                "n": int(row[n_col]),
                "section": row[sec_col],
                "subsection": row[sub_col] if sub_col is not None else None,
                "question": row[q_col],
                "answer_type": row[t_col],
                "example": row[ex_col],
                "key": row[key_col],
            })
        return out

    p1 = _extract("Phase 1 - ONBOARDING", phase=1,
                  n_col=0, sec_col=1, sub_col=None, q_col=2, t_col=3, ex_col=4, key_col=5,
                  header_row=1)
    p2 = _extract("Phase 2 - Sales Book", phase=2,
                  n_col=0, sec_col=1, sub_col=2, q_col=3, t_col=4, ex_col=5, key_col=6,
                  header_row=1)
    p3 = _extract("Phase 3 - VP Conversion", phase=3,
                  n_col=0, sec_col=1, sub_col=None, q_col=3, t_col=4, ex_col=5, key_col=2,
                  header_row=2)

    expected = (57, 22, 35)
    actual = (len(p1), len(p2), len(p3))
    if actual != expected:
        print(f"WARNING: counts mismatch — got {actual}, expected {expected}", file=sys.stderr)

    return p1 + p2 + p3


def write_outputs(all_q: list[dict]) -> None:
    DOMAIN_DIR.mkdir(parents=True, exist_ok=True)
    DATA_OUT.write_text(json.dumps(all_q, indent=2, default=str), encoding="utf-8")

    py_source = '''"""Auto-generated onboarding registry. /Oliviercontribution.

DO NOT edit by hand. Regenerate via:
    python3 backend/scripts/generate_onboarding_registry.py

Source: Google Sheet ''' + SHEET_ID + '''
"""

from __future__ import annotations

import json
from pathlib import Path

_DATA: list[dict] = json.loads(
    (Path(__file__).parent / "_onboarding_registry.json").read_text(encoding="utf-8")
)

ONBOARDING_QUESTIONS: dict[str, dict] = {q["key"]: q for q in _DATA}


def get_phase_questions(phase: int) -> dict[str, dict]:
    return {k: v for k, v in ONBOARDING_QUESTIONS.items() if v.get("phase") == phase}


def get_section_questions(phase: int, section: str) -> dict[str, dict]:
    return {
        k: v for k, v in ONBOARDING_QUESTIONS.items()
        if v.get("phase") == phase and v.get("section") == section
    }


def get_phase_sections(phase: int) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for v in ONBOARDING_QUESTIONS.values():
        if v.get("phase") == phase and v.get("section") not in seen:
            seen.add(v["section"])
            out.append(v["section"])
    return out


PHASE_1_TOTAL = len(get_phase_questions(1))
PHASE_2_TOTAL = len(get_phase_questions(2))
PHASE_3_TOTAL = len(get_phase_questions(3))
TOTAL_QUESTIONS = PHASE_1_TOTAL + PHASE_2_TOTAL + PHASE_3_TOTAL

assert TOTAL_QUESTIONS == 114, f"Expected 114 questions, got {TOTAL_QUESTIONS}"


__all__ = [
    "ONBOARDING_QUESTIONS",
    "get_phase_questions",
    "get_section_questions",
    "get_phase_sections",
    "PHASE_1_TOTAL",
    "PHASE_2_TOTAL",
    "PHASE_3_TOTAL",
    "TOTAL_QUESTIONS",
]
'''
    PY_OUT.write_text(py_source, encoding="utf-8")
    print(f"[generate-onboarding-registry] wrote {DATA_OUT}")
    print(f"[generate-onboarding-registry] wrote {PY_OUT}")
    print(f"[generate-onboarding-registry] {len(all_q)} questions total")


def main() -> int:
    fetch_sheet()
    questions = extract()
    write_outputs(questions)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
